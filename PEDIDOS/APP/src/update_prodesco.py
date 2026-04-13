import json
import openpyxl

EXCEL_PATH = r'C:\Users\jmgon\Desktop\Claude Code\Casa Amparo 1948\PEDIDOS\APP\docs\Base Productos Casa Amparo.xlsx'
JSON_PATH = r'C:\Users\jmgon\Desktop\Claude Code\Casa Amparo 1948\PEDIDOS\APP\src\productos.json'

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['Productos']

# Find last Prodesco row
last_prodesco_row = None
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value == 'Prodesco':
        last_prodesco_row = row[0].row

next_row = last_prodesco_row + 1

new_products = [
    # KETCHUP
    ("SALSAS Y CONDIMENTOS", "4648", "Ketchup Bag in Box 2.75kg (Helios)", 11.45, "ud", "10%"),
    ("SALSAS Y CONDIMENTOS", "3547", "Ketchup Bote 1.85kg (AMH)", 3.75, "ud", "10%"),
    ("SALSAS Y CONDIMENTOS", "5631", "Ketchup Sobres 10g Estuche 275ud (AMH)", 9.58, "caja", "10%"),
    ("SALSAS Y CONDIMENTOS", "9414", "Ketchup Tarrito Cristal 30g Caja 24ud (Helios)", 7.14, "caja", "10%"),
    # MOSTAZA
    ("SALSAS Y CONDIMENTOS", "4621", "Mostaza Antigua Grano Cubo 1kg (Bornier)", 6.35, "ud", "10%"),  # ya existe, skip
    ("SALSAS Y CONDIMENTOS", "4625", "Mostaza Original Dijon Cubo 1kg (Bornier)", 6.82, "ud", "10%"),
    ("SALSAS Y CONDIMENTOS", "3671", "Mostaza Bote 1.85kg (AMH)", 3.80, "ud", "10%"),
    ("SALSAS Y CONDIMENTOS", "4840", "Mostaza Sobres 6g Caja 300ud (AMH)", 8.59, "caja", "10%"),
    ("SALSAS Y CONDIMENTOS", "7125", "Mostaza Sobres 10g Estuche 200ud (Heinz)", 11.72, "caja", "10%"),
    ("SALSAS Y CONDIMENTOS", "8605", "Mostaza Tarrito Cristal 28g Caja 24ud (Kuhne)", 7.28, "caja", "10%"),
    # MAHONESA MONODOSIS
    ("SALSAS Y CONDIMENTOS", "9416", "Mahonesa Sobres 10g Estuche 200ud (Heinz)", 13.45, "caja", "10%"),
    ("SALSAS Y CONDIMENTOS", "8219", "Mahonesa Tarrito Cristal 25g Caja 24ud (Heinz)", 7.09, "caja", "10%"),
    # ACEITE OLIVA VIRGEN EXTRA
    ("ACEITES Y VINAGRES", "8645", "Aceite Oliva Virgen Extra Garrafa 5L (AMH)", 28.99, "ud", "10%"),
    ("ACEITES Y VINAGRES", "10029", "Aceite Oliva Virgen Extra Garrafa 5L (Hacholiva)", 27.99, "ud", "10%"),
    ("ACEITES Y VINAGRES", "8681", "Aceite Oliva Virgen Extra Botella 250ml (La Flor de Malaga)", 2.28, "ud", "10%"),
    # ACEITE OLIVA SUAVE (FREIR)
    ("ACEITES Y VINAGRES", "1071", "Aceite Oliva Suave Garrafa 5L (AMH)", 22.69, "ud", "10%"),
    # ACEITE GIRASOL
    ("ACEITES Y VINAGRES", "3001", "Aceite Girasol Garrafa 5L (AMH)", 9.99, "ud", "10%"),
    ("ACEITES Y VINAGRES", "3000", "Aceite Girasol Bidon 25L (AMH)", 46.99, "ud", "10%"),
    ("ACEITES Y VINAGRES", "2268", "Aceite Girasol Alto Oleico 80% Garrafa 5L (AMH)", 11.45, "ud", "10%"),
    # ACEITE MONODOSIS
    ("ACEITES Y VINAGRES", "8780", "Aceite Oliva VE Tarrina Monodosis 10ml Caja 240ud (AMH)", 20.19, "caja", "10%"),
    ("ACEITES Y VINAGRES", "8877", "Aceite Oliva VE Tarrina Monodosis 10ml Caja 240ud (La Flor de Malaga)", 19.99, "caja", "10%"),
    # SAL
    ("SALSAS Y CONDIMENTOS", "6047", "Sal Escamas Maldon Natural Caja 250g", 5.16, "ud", "10%"),
    ("SALSAS Y CONDIMENTOS", "6431", "Sal Escamas Maldon Natural Caja 125g", 3.49, "ud", "10%"),
    ("SALSAS Y CONDIMENTOS", "8736", "Sal Escamas Maldon Natural Cubo 1.4kg", 21.35, "ud", "10%"),
    ("SALSAS Y CONDIMENTOS", "2920", "Sal Fina Seca Marina Bolsa 1kg (Leda)", 0.30, "ud", "10%"),
    # PIMIENTA
    ("ESPECIAS", "3249", "Pimienta Negra Grano Tarro 700g (AMH)", 10.15, "ud", "10%"),
    ("ESPECIAS", "3252", "Pimienta Negra Molida Tarro 800g (AMH)", 11.29, "ud", "10%"),
    ("ESPECIAS", "2895", "Pimienta Blanca Grano Tarro 800g (AMH)", 8.65, "ud", "10%"),
    ("ESPECIAS", "3248", "Pimienta Blanca Molida Tarro 800g (AMH)", 14.27, "ud", "10%"),
    # PIMENTON
    ("ESPECIAS", "2901", "Pimenton Extra Dulce Tarro 800g (AMH)", 5.75, "ud", "10%"),
    ("ESPECIAS", "2902", "Pimenton Dulce Ahumado DO Lata 750g (La Vera)", 11.39, "ud", "10%"),
    ("ESPECIAS", "7534", "Pimenton Picante Ahumado DO Lata 750g (La Vera)", 12.97, "ud", "10%"),
    # OREGANO
    ("ESPECIAS", "3223", "Oregano Hojas Tarro 130g (AMH)", 2.59, "ud", "10%"),
    # COMINO
    ("ESPECIAS", "2891", "Comino Molido Tarro 700g (Maripaz)", 6.65, "ud", "10%"),
    # LAUREL
    ("ESPECIAS", "2883", "Laurel Hojas Tarro 80g (AMH)", 1.90, "ud", "10%"),
    # CANELA
    ("ESPECIAS", "3329", "Canela Molida Tarro 600g (AMH)", 4.69, "ud", "10%"),
    # NUEZ MOSCADA
    ("ESPECIAS", "2898", "Nuez Moscada Molida Tarro 900g (AMH)", 20.09, "ud", "10%"),
    # CAYENA
    ("ESPECIAS", "2897", "Cayena Molida (Chile) Tarro 900g (AMH)", 7.45, "ud", "10%"),
    # PEREJIL SECO
    ("ESPECIAS", "3644", "Perejil Hoja Tarro 130g (AMH)", 3.13, "ud", "10%"),
    # AJO GRANULADO
    ("ESPECIAS", "4897", "Ajo Granulado Tarro 900g (AMH)", 7.30, "ud", "10%"),
    # TOMILLO
    ("ESPECIAS", "3537", "Tomillo Hojas Tarro 360g (AMH)", 2.85, "ud", "10%"),
    # ROMERO
    ("ESPECIAS", "4325", "Romero Hojas Tarro 330g (AMH)", 3.05, "ud", "10%"),
    # VINAGRE
    ("ACEITES Y VINAGRES", "5404", "Vinagre de Jerez Garrafa 5L (Riojavina)", 9.48, "ud", "10%"),
    ("ACEITES Y VINAGRES", "8949", "Vinagre de Modena Balsamico IGP Garrafa 5L (Lunios)", 10.55, "ud", "10%"),
    ("ACEITES Y VINAGRES", "4847", "Vinagre de Vino Blanco Garrafa 5L (Merry)", 3.91, "ud", "10%"),
    ("ACEITES Y VINAGRES", "6420", "Vinagre de Vino Tinto Garrafa 5L (AMH)", 3.91, "ud", "10%"),
    ("ACEITES Y VINAGRES", "8784", "Vinagre de Vino Tarrina Monodosis 10ml Caja 240ud (La Flor de Malaga)", 12.05, "caja", "10%"),
    # PULPO
    ("REFRIGERADO", "716", "Pulpo T3 Cocido 1 Pata Skin Bandeja 180-200g (Nuchar)", 9.65, "ud", "10%"),
    ("REFRIGERADO", "589", "Pulpo T4 Cocido 2 Patas Skin Bandeja 200g (Nuchar)", 8.35, "ud", "10%"),
    ("REFRIGERADO", "1556", "Pulpo T4 Cocido en su Jugo 1 Pata Skin Bandeja 150-200g (Nuchar)", 7.15, "ud", "10%"),
    ("REFRIGERADO", "4881", "Pulpo T2 1 Pata Cocido en su Jugo Bolsa 300-400g (LFP)", 32.90, "kg", "10%"),
]

# Check which codes already exist for Prodesco
existing_codes = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value == 'Prodesco':
        code = str(row[2].value) if row[2].value else ''
        existing_codes.add(code)

added = 0
skipped = 0
# Insert new rows after the last Prodesco row
insert_at = next_row
for cat, codigo, nombre, precio, unidad, iva in new_products:
    if codigo in existing_codes:
        print(f"  SKIP (ya existe): {codigo} - {nombre}")
        skipped += 1
        continue
    ws.insert_rows(insert_at)
    ws.cell(row=insert_at, column=1, value='Prodesco')
    ws.cell(row=insert_at, column=2, value=cat)
    ws.cell(row=insert_at, column=3, value=codigo)
    ws.cell(row=insert_at, column=4, value=nombre)
    ws.cell(row=insert_at, column=5, value=precio)
    ws.cell(row=insert_at, column=6, value=unidad)
    ws.cell(row=insert_at, column=7, value=iva)
    insert_at += 1
    added += 1
    existing_codes.add(codigo)

print(f"\nAnadidos: {added}, Saltados (duplicados): {skipped}")

wb.save(EXCEL_PATH)
print("Excel guardado")

# Regenerate JSON
wb2 = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws2 = wb2['Productos']
productos = []
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
    proveedor, categoria, codigo, producto, precio, unidad, iva = row
    if not proveedor or not producto:
        continue
    productos.append({
        'proveedor': proveedor,
        'categoria': categoria or '',
        'codigo': str(codigo) if codigo else None,
        'producto': producto,
        'precio': precio if isinstance(precio, (int, float)) else precio,
        'unidad': str(unidad) if unidad else 'ud',
        'iva': str(iva) if iva else '10%'
    })

with open(JSON_PATH, 'w', encoding='utf-8') as f:
    json.dump(productos, f, ensure_ascii=False, indent=2)

prodesco = [p for p in productos if p['proveedor'] == 'Prodesco']
print(f"\nJSON regenerado: {len(productos)} productos totales")
print(f"Prodesco ahora: {len(prodesco)} productos")

# Show Prodesco by category
cats = {}
for p in prodesco:
    cats.setdefault(p['categoria'], []).append(p)
for cat, items in sorted(cats.items()):
    print(f"\n  {cat} ({len(items)}):")
    for p in items:
        print(f"    {p['codigo']} - {p['producto']} - {p['precio']} EUR/{p['unidad']}")
