"""
excel_to_json.py — Extrae fichas técnicas del Excel "FT Casa Amparo v2.xlsx"
y genera data/recetas.json + data/fotos/<slug>.<ext> para la app del recetario.

NO incluye datos económicos (costes, PVP, márgenes). Cocina nunca los ve.

Uso:
    python tools/excel_to_json.py [ruta_al_excel]
Por defecto lee H:/Mi unidad/02_Operaciones/FT Casa Amparo v2.xlsx
"""

import json
import os
import re
import sys
import shutil
import zipfile
import unicodedata
import warnings
import io
from datetime import datetime
from xml.etree import ElementTree as ET

warnings.filterwarnings("ignore")
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl  # noqa: E402

DEFAULT_EXCEL = r"H:/Mi unidad/02_Operaciones/FT Casa Amparo v2.xlsx"
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
FOTOS_DIR = os.path.join(DATA_DIR, "fotos")
JSON_PATH = os.path.join(DATA_DIR, "recetas.json")

SKIP_SHEETS = {
    "(aaa) BBDD Ingredientes",
    "Carta CCA1948",
    "Carta Menu CCA1948",
    "Carta Entrehoras CCA1948",
    "FT Platos CCA1948",
    "Comparativa FT vs Cartas",
}

# Overrides explícitos: hoja -> nombre exacto de plato en FT Platos CCA1948.
# Usar cuando el matcher no acierta por singular/plural u otra ambigüedad lingüística.
OVERRIDES = {
    "ZAMBURIÑAS PICANTONAS": "Zamboriñas a la bilbaína",
    "BOCATA CHIPIRONES": "Bocadillo de chipironcitos a la andaluza",
    "BOCATA LOMO CHEDDAR": "Bocadillo de lomo adobado con queso",
    "BOCATA PIRIPI": "Bocadillo Piripi",
    "GAMBON AJILLO": "Gambones al pil pil con tomate cherry y Jerez",
    "PATA PULPO MOJO VERDE": "Pulpo a la plancha con mojo verde de cilantro",
    "PAN DE LA CASA PULPO": "Pan de queso tetilla y pulpo",
    "CHIPIRONCITOS HUEVO FRITO": "Huevos rotos con chipironcitos a la andaluza",
    "MACARRONES AMPARO": "Los macarrones de Amparo (con champiñones y chorizo de León)",
    "MACARRONES INFANTILES TOM": "Pasta con tomate (guarnición infantil)",
    "TOMATE TEMPORADA BONITO": "Tomates aliñados con bonito, encurtidos y aguacate",
    "MN PATA PULPO REVOLCONA": "Pulpo a la plancha con mojo verde (versión menú, supl. +4€)",
    "MNCALLOS GARBANZOS": "La cuchara del día",
    "MN LENTEJAS TRADICIONALES": "La cuchara del día",
    "MN POCHAS ALMEJAS": "La cuchara del día",
    "MN POCHAS GAMBONES": "La cuchara del día",
    "MN GARBANZOS PULPO": "La cuchara del día",
    "ALBONDIGAS PATATAS FRITAS": "Albóndigas de ternera en salsa",
    "TARTAR ATUN HUEVO FRITO": "Huevos rotos con dados de atún rojo y trufa",
}


def is_menu_sheet(sheet_name):
    """Detecta hojas del menú: tanto con espacio (MN ALCACHOFA) como pegado (MNCALLOS)."""
    s = sheet_name.upper().strip()
    return s.startswith("MN ") or (s.startswith("MN") and len(s) > 2 and s[2].isalpha())


def slugify(text):
    text = unicodedata.normalize("NFKD", str(text))
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text


def split_elaboracion(raw):
    """Divide el texto de elaboración en pre / elaboracion / emplatado."""
    if not raw:
        return {"pre": "", "elaboracion": "", "emplatado": "", "raw": ""}
    text = str(raw).replace("\r\n", "\n").strip()
    sections = {"pre": "", "elaboracion": "", "emplatado": ""}
    pattern = re.compile(
        r"(PRE\s*ELABORACION|PREELABORACION|ELABORACION|EMPLATAMOS|EMPLATADO|MONTAJE)",
        re.IGNORECASE,
    )
    parts = pattern.split(text)
    if len(parts) <= 1:
        return {"pre": "", "elaboracion": text, "emplatado": "", "raw": text}
    for i in range(1, len(parts), 2):
        marker = parts[i].upper().replace(" ", "")
        body = parts[i + 1].strip() if i + 1 < len(parts) else ""
        body = body.lstrip(":\n ").strip()
        if marker.startswith("PRE"):
            sections["pre"] = body
        elif marker.startswith("EMPLAT") or marker.startswith("MONTAJE"):
            sections["emplatado"] = body
        else:
            sections["elaboracion"] = body
    sections["raw"] = text
    return sections


def parse_ingredientes_principales(rows):
    """Lee ingredientes en columnas A–D (índices 0..3) entre filas 6 y 19."""
    ingredientes = []
    for row in rows[5:19]:
        if not row:
            continue
        nombre = row[0] if len(row) > 0 else None
        med = row[1] if len(row) > 1 else None
        udm = row[2] if len(row) > 2 else None
        notas = row[3] if len(row) > 3 else None
        if nombre is None or str(nombre).strip() == "":
            continue
        if str(nombre).strip() in {"0", "—"}:
            continue
        ingredientes.append(
            {
                "nombre": str(nombre).strip(),
                "med": med if med not in (None, "") else None,
                "udm": str(udm).strip() if udm not in (None, "") else "",
                "notas": str(notas).strip() if notas not in (None, "") else "",
            }
        )
    return ingredientes


def parse_alergenos(rows):
    """Recoge los alérgenos del espejo de coste (col T, índice 19)."""
    alergenos = set()
    for row in rows[4:19]:
        if not row or len(row) <= 19:
            continue
        val = row[19]
        if val and isinstance(val, str) and val.strip():
            for a in val.split(","):
                a = a.strip()
                if a:
                    alergenos.add(a)
    return sorted(alergenos)


def parse_pcg(rows):
    """PCG suele estar en torno a la fila 41–42, columna A."""
    for i in range(35, min(50, len(rows))):
        row = rows[i]
        if not row:
            continue
        cell = row[0] if len(row) > 0 else None
        if cell and isinstance(cell, str) and cell.strip().startswith("P.C.G"):
            # texto en la fila siguiente
            if i + 1 < len(rows):
                next_cell = rows[i + 1][0] if rows[i + 1] else None
                if next_cell:
                    return [
                        line.strip().lstrip("•").strip()
                        for line in str(next_cell).split("\n")
                        if line.strip()
                    ]
    return []


def parse_subrecetas(rows):
    """Detecta secciones SUBRECETA: <nombre> (codigo) y extrae sus ingredientes y elaboración."""
    subrecetas = []
    i = 50  # suelen empezar en la fila 51
    while i < len(rows):
        row = rows[i]
        cell = row[0] if row and len(row) > 0 else None
        if cell and isinstance(cell, str) and cell.strip().upper().startswith("SUBRECETA"):
            header = cell.strip()
            m = re.match(r"SUBRECETA:\s*(.+?)\s*\(([^)]+)\)", header, re.IGNORECASE)
            nombre = m.group(1).strip() if m else header.replace("SUBRECETA:", "").strip()
            codigo = m.group(2).strip() if m else ""
            elaboracion = ""
            ingredientes = []
            j = i + 2  # saltar fila de cabecera
            first_elab = True
            while j < len(rows):
                rr = rows[j]
                if not rr:
                    j += 1
                    continue
                first = rr[0] if len(rr) > 0 else None
                if first and isinstance(first, str) and first.strip().upper().startswith("SUBRECETA"):
                    break
                if first is None or str(first).strip() == "":
                    j += 1
                    continue
                med = rr[1] if len(rr) > 1 else None
                udm = rr[2] if len(rr) > 2 else None
                elab_cell = rr[3] if len(rr) > 3 else None
                ingredientes.append(
                    {
                        "nombre": str(first).strip(),
                        "med": med if med not in (None, "") else None,
                        "udm": str(udm).strip() if udm not in (None, "") else "",
                    }
                )
                if first_elab and elab_cell and isinstance(elab_cell, str) and elab_cell.strip():
                    elaboracion = elab_cell.strip()
                    first_elab = False
                j += 1
            subrecetas.append(
                {
                    "nombre": nombre,
                    "codigo": codigo,
                    "ingredientes": ingredientes,
                    "elaboracion": elaboracion,
                }
            )
            i = j
        else:
            i += 1
    return subrecetas


def build_sheet_to_images(excel_path):
    """Mapea nombre de hoja → lista de paths internos (xl/media/imageX.ext)."""
    BS = chr(92)
    with zipfile.ZipFile(excel_path) as z:
        wb_rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        sheet_targets = {r.get("Id"): r.get("Target") for r in wb_rels}
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        main_ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        rid_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
        sheets = []
        for s in wb.find(main_ns + "sheets"):
            sheets.append((s.get("name"), s.get(rid_ns + "id")))
        result = {}
        for name, rid in sheets:
            target = sheet_targets[rid]
            sheet_xml = ("xl/" + target).replace(BS, "/")
            d, f = os.path.split(sheet_xml)
            rels_path = d + "/_rels/" + f + ".rels"
            if rels_path not in z.namelist():
                result[name] = []
                continue
            rels = ET.fromstring(z.read(rels_path))
            drawings = [r.get("Target") for r in rels if "drawing" in r.get("Type", "")]
            imgs = []
            for dt in drawings:
                dpath = os.path.normpath(os.path.join(d, dt)).replace(BS, "/")
                ddir, dfile = os.path.split(dpath)
                drels = ddir + "/_rels/" + dfile + ".rels"
                if drels not in z.namelist():
                    continue
                dr = ET.fromstring(z.read(drels))
                for r in dr:
                    if "image" in r.get("Type", ""):
                        img_t = r.get("Target")
                        img_path = os.path.normpath(os.path.join(ddir, img_t)).replace(BS, "/")
                        imgs.append(img_path)
            result[name] = imgs
    return result


def extract_image(excel_path, internal_path, dest_path):
    with zipfile.ZipFile(excel_path) as z:
        with z.open(internal_path) as src, open(dest_path, "wb") as dst:
            shutil.copyfileobj(src, dst)


def normalize_text(t):
    t = unicodedata.normalize("NFKD", str(t))
    t = t.encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"[^a-z0-9 ]", " ", t)


STOP = {
    "de", "la", "el", "los", "las", "con", "y", "o", "a", "al", "del",
    "en", "un", "una", "su", "sus", "para", "por",
}


def tokens(t):
    return [w for w in normalize_text(t).split() if w and w not in STOP and len(w) > 2]


def best_match(sheet_name, sheet_title, candidates):
    """Devuelve el plato del índice que mejor encaja con la hoja.

    Reglas:
    1. Construye el set de tokens del título y nombre de la hoja.
    2. Exige que el PRIMER token significativo del título de la hoja aparezca en el candidato.
    3. Si la hoja empieza por 'MN ', solo considera candidatos del menú (carta == 'menu').
    4. Entre los que pasan filtros, ranking por Jaccard (overlap/union) con desempate por overlap.
    5. Si no hay candidato que pase, devuelve None (mejor sin match que match incorrecto).
    """
    title_tokens = tokens(sheet_title)
    name_tokens = tokens(sheet_name)
    sheet_tokens = set(title_tokens) | set(name_tokens)
    if not sheet_tokens or not title_tokens:
        return None

    first_token = title_tokens[0]
    menu_sheet = is_menu_sheet(sheet_name)

    best = None
    best_score = 0.0
    best_overlap = 0
    for c in candidates:
        cartas = c.get("cartas") or []
        if menu_sheet:
            if "menu" not in cartas:
                continue
        else:
            # Hojas no-MN: excluir candidatos que SOLO están en el menú
            if cartas == ["menu"]:
                continue
        ct = set(tokens(c["plato"]))
        if not ct or first_token not in ct:
            continue
        overlap = len(sheet_tokens & ct)
        union = len(sheet_tokens | ct)
        jaccard = overlap / union if union else 0
        if jaccard > best_score or (jaccard == best_score and overlap > best_overlap):
            best_score = jaccard
            best_overlap = overlap
            best = c
    return best


def parse_carta_index(wb):
    """Lee FT Platos CCA1948 -> [{plato, seccion, cartas[]}]."""
    ws = wb["FT Platos CCA1948"]
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:
            continue
        if not row or row[0] is None:
            continue
        try:
            num = int(row[0])
        except (ValueError, TypeError):
            continue
        seccion = row[1] or ""
        plato = row[2] or ""
        cartas = []
        if row[3]:
            cartas.append("principal")
        if row[4]:
            cartas.append("entrehoras")
        if row[5]:
            cartas.append("menu")
        seccion_str = str(seccion).strip()
        plato_str = str(plato).strip()
        # Platos infantiles (sección "Para los peques" o nombre con "(infantil)") van
        # a su propia carta "infantil" en lugar de aparecer en la carta principal.
        is_infantil = (
            "peques" in seccion_str.lower()
            or "infantil" in plato_str.lower()
        )
        if is_infantil:
            cartas = ["infantil"]
        out.append(
            {
                "num": num,
                "plato": plato_str,
                "seccion": seccion_str,
                "cartas": cartas,
            }
        )
    return out


def find_description(nombre_carta, descripciones):
    """Busca la descripción comercial. Igualdad normalizada y luego fuzzy por Jaccard."""
    if not nombre_carta:
        return ""
    key = normalize_text(nombre_carta)
    if key in descripciones:
        return descripciones[key]
    target = set(tokens(nombre_carta))
    if not target:
        return ""
    best = ""
    best_score = 0.0
    for k, v in descripciones.items():
        kt = set(tokens(k))
        ov = len(target & kt)
        if ov < 2:
            continue
        union = len(target | kt)
        score = ov / union if union else 0
        if score > best_score:
            best_score = score
            best = v
    # exigir un mínimo de calidad (>0.5 Jaccard) para evitar descripciones disparatadas
    return best if best_score >= 0.5 else ""


def parse_carta_descripciones(wb):
    """Lee las 3 cartas y devuelve {plato_normalizado: descripcion}."""
    descripciones = {}
    for sheet in ["Carta CCA1948", "Carta Menu CCA1948", "Carta Entrehoras CCA1948"]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            # Las filas válidas de plato terminan en un número (precio)
            has_price = any(isinstance(v, (int, float)) for v in row)
            if not has_price:
                continue
            strs = [str(v).strip() for v in row if isinstance(v, str) and v.strip()]
            if len(strs) < 2:
                continue
            # Plato = penúltimo texto, descripción = último texto.
            # (Si sólo hay 2 textos, primero es plato y segundo descripción.)
            plato = strs[-2]
            desc = strs[-1]
            # Filtrar encabezados de sección (todo mayúsculas y muy cortos)
            if plato.isupper() and len(plato) < 25:
                continue
            if len(desc) > 15:
                descripciones[normalize_text(plato)] = desc
    return descripciones


def parse_recipe_sheet(ws, sheet_name, image_paths, excel_path, fotos_dir):
    rows = list(ws.iter_rows(values_only=True))
    # Título: row 3, col A
    titulo = ""
    if len(rows) > 2 and rows[2]:
        titulo = str(rows[2][0] or "").strip()
    if not titulo:
        titulo = sheet_name
    # Elaboración: row 6, col E (índice 4)
    elab_raw = ""
    if len(rows) > 5 and rows[5] and len(rows[5]) > 4:
        elab_raw = rows[5][4] or ""
    elab = split_elaboracion(elab_raw)
    ingredientes = parse_ingredientes_principales(rows)
    alergenos = parse_alergenos(rows)
    pcg = parse_pcg(rows)
    subrecetas = parse_subrecetas(rows)

    # Foto
    foto_rel = ""
    if image_paths:
        internal = image_paths[0]
        ext = os.path.splitext(internal)[1].lower()
        slug = slugify(sheet_name)
        dest_name = f"{slug}{ext}"
        dest = os.path.join(fotos_dir, dest_name)
        extract_image(excel_path, internal, dest)
        foto_rel = f"data/fotos/{dest_name}"

    return {
        "id": slugify(sheet_name),
        "sheet": sheet_name,
        "nombre": titulo,
        "ingredientes": ingredientes,
        "elaboracion": elab,
        "alergenos": alergenos,
        "pcg": pcg,
        "subrecetas": subrecetas,
        "foto": foto_rel,
    }


def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL
    if not os.path.exists(excel_path):
        print(f"ERROR: no se encuentra el Excel en {excel_path}")
        sys.exit(1)

    os.makedirs(FOTOS_DIR, exist_ok=True)
    # limpiar fotos previas
    for f in os.listdir(FOTOS_DIR):
        os.remove(os.path.join(FOTOS_DIR, f))

    print(f"Leyendo {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    sheet_to_images = build_sheet_to_images(excel_path)
    indice = parse_carta_index(wb)
    descripciones = parse_carta_descripciones(wb)

    recetas = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        try:
            r = parse_recipe_sheet(
                ws, sheet_name, sheet_to_images.get(sheet_name, []), excel_path, FOTOS_DIR
            )
        except Exception as e:
            print(f"  ERROR parseando {sheet_name}: {e}")
            continue
        # cruzar con índice carta. Probamos override por nombre exacto y por trim
        # (algunas hojas tienen espacios al final accidentalmente).
        match = None
        for key in (sheet_name, sheet_name.strip(), sheet_name.strip().upper()):
            if key in OVERRIDES:
                target = OVERRIDES[key]
                target_norm = normalize_text(target)
                for c in indice:
                    if normalize_text(c["plato"]) == target_norm:
                        match = c
                        break
                if match:
                    break
        if match is None:
            match = best_match(sheet_name, r["nombre"], indice)
        if match:
            r["nombre_carta"] = match["plato"].replace("Zamboriñas", "Zamburiñas")
            r["seccion"] = match["seccion"]
            r["cartas"] = match["cartas"]
        else:
            r["nombre_carta"] = r["nombre"]
            r["seccion"] = ""
            r["cartas"] = []
        # Fallback: cualquier hoja con prefijo MN se etiqueta como menú aunque
        # no aparezca en el índice (postres del menú, platos nuevos, etc.).
        if is_menu_sheet(sheet_name) and "menu" not in r["cartas"]:
            r["cartas"] = (r["cartas"] or []) + ["menu"]
        # descripción comercial si existe
        desc = find_description(r["nombre_carta"], descripciones)
        r["descripcion"] = desc
        recetas.append(r)
        ico = "[F]" if r["foto"] else "[ ]"
        print(f"  {ico} {sheet_name:<35} -> {len(r['ingredientes'])} ing - {len(r['subrecetas'])} subr - {r['seccion']}")

    secciones = sorted({r["seccion"] for r in recetas if r["seccion"]})

    output = {
        "generado": datetime.now().isoformat(timespec="seconds"),
        "version": 1,
        "total": len(recetas),
        "secciones": secciones,
        "cartas": ["principal", "entrehoras", "menu", "infantil"],
        "recetas": sorted(recetas, key=lambda x: (x["seccion"], x["nombre"])),
    }

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    fotos_count = sum(1 for r in recetas if r["foto"])
    print(f"\nOK - {len(recetas)} recetas escritas en {JSON_PATH}")
    print(f"OK - {fotos_count} fotos extraidas a {FOTOS_DIR}")
    print(f"OK - {len(secciones)} secciones, {len({s for r in recetas for s in r['cartas']})} cartas")


if __name__ == "__main__":
    main()
