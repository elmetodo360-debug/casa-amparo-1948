"""
excel_to_json.py — Convierte el Excel maestro de la Comida de Familia
en `data/menu.json` consumible por la SPA.

Lee `H:\\Mi unidad\\02_Operaciones\\Comida de Familia\\Comida de Familia.xlsx`
(generado a su vez por `generar_excel_comida_familia.py`).

Estructura de salida (data/menu.json):
{
  "actualizado": "2026-04-29",
  "personas": 10,
  "normas": [{"titulo": ..., "detalle": ...}, ...],
  "calendario": [
      {"semana": 1, "dia": "Lunes", "plato": "...", "guarnicion": "...", "esPescado": false}, ...
  ],
  "recetas": [
      {"semana": 1, "dia": "Lunes", "plato": "...", "esPescado": false,
       "guarnicion": "...", "pasos": ["...", "..."],
       "ingredientes": [{"tipo": "Proteína", "nombre": "...", "proveedor": "...",
                         "codigo": "...", "udm": "GR", "cantidad": 2000}, ...]},
      ...
  ],
  "compraTotal": [{"proveedor": "...", "codigo": "...", "producto": "...",
                   "udm": "GR", "cantidad": 4000}, ...]
}

NUNCA se incluyen €, costes ni precios. La app es para personal/cocina, no
deben ver datos económicos.

Uso:
    python tools/excel_to_json.py
"""
import io
import json
import os
import sys
import warnings
from datetime import date

warnings.filterwarnings("ignore")
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl  # noqa: E402

EXCEL_PATH = r"H:\Mi unidad\02_Operaciones\Comida de Familia\Comida de Familia.xlsx"
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSON_PATH = os.path.join(PROJECT_ROOT, "data", "menu.json")


def leer_normas(wb):
    ws = wb["Normas"]
    normas = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        # cols B..F: # | Norma | Detalle (D-F merged)
        if not row or row[1] is None or not isinstance(row[1], int):
            # quizás fila de horarios o vacía → parar al detectar
            if row and row[1] in ("HORARIOS DE COMIDA DE PERSONAL", "Mañana", "Tarde"):
                continue
            if row and row[1] is None:
                continue
            # detener en cuanto B no es int contiguo
            break
        titulo = row[2] or ""
        detalle = row[3] or ""
        normas.append({"numero": row[1], "titulo": str(titulo).strip(), "detalle": str(detalle).strip()})
    return normas


def leer_menu(wb):
    """Lee el calendario desde la hoja 'Compra Detalle' (más fiable que la hoja
    Calendario porque no tiene celdas merged complejas).
    """
    ws = wb["Compra Detalle"]
    calendario_set = {}
    recetas_data = {}  # (sem, dia) -> {plato, guarnicion, esPescado, ingredientes:[...]}
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or not row[1]:
            continue
        sem_str, dia, plato, tipo, ingrediente, prov, cod, udm, cantidad = row[1:10]
        if not sem_str or not str(sem_str).startswith("S"):
            continue
        try:
            sem = int(str(sem_str)[1:])
        except ValueError:
            continue
        es_pescado = "🐟" in (plato or "")
        plato_clean = str(plato or "").replace("🐟 ", "")
        key = (sem, dia)
        if key not in recetas_data:
            recetas_data[key] = {
                "semana": sem, "dia": dia, "plato": plato_clean,
                "esPescado": es_pescado, "guarnicion": "",
                "ingredientes": [],
            }
        recetas_data[key]["ingredientes"].append({
            "tipo": tipo, "nombre": str(ingrediente or ""),
            "proveedor": str(prov or ""), "codigo": str(cod or ""),
            "udm": str(udm or ""), "cantidad": cantidad,
        })

    # guarnición = mirar ingredientes tipo Guarnición y deducir el nombre genérico
    guarn_map = {
        "ARROZ EXTRA": "Arroz blanco",
        "PATATA AGRIA": "Patatas fritas",
        "LECHUGA ICEBERG": "Ensalada mixta",
    }
    for key, data in recetas_data.items():
        for ing in data["ingredientes"]:
            if ing["tipo"] == "Guarnición" and ing["nombre"] in guarn_map:
                data["guarnicion"] = guarn_map[ing["nombre"]]
                break

    # Calendario simplificado
    DIA_ORDEN = {"Lunes": 1, "Martes": 2, "Miércoles": 3, "Jueves": 4, "Viernes": 5, "Sábado": 6}
    cal = sorted(
        [{
            "semana": d["semana"], "dia": d["dia"], "plato": d["plato"],
            "guarnicion": d["guarnicion"], "esPescado": d["esPescado"],
        } for d in recetas_data.values()],
        key=lambda x: (x["semana"], DIA_ORDEN.get(x["dia"], 99)),
    )

    recetas = sorted(recetas_data.values(), key=lambda x: (x["semana"], DIA_ORDEN.get(x["dia"], 99)))
    return cal, recetas


def leer_recetas_pasos(wb):
    """Lee la hoja Recetario para extraer pasos. Devuelve {plato_normalizado: [pasos]}."""
    ws = wb["Recetario"]
    pasos_por_plato = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or not row[1]:
            continue
        sem_str, dia, plato, pasos_txt = row[1], row[2], row[3], row[4]
        if not plato or not pasos_txt:
            continue
        plato_norm = str(plato).replace("🐟 ", "").strip()
        # parsear pasos numerados
        lineas = [ln.strip() for ln in str(pasos_txt).split("\n") if ln.strip()]
        pasos = []
        for ln in lineas:
            if ln.startswith("Guarnición:"):
                continue
            # quitar prefijo numérico
            if "." in ln:
                _, _, resto = ln.partition(".")
                pasos.append(resto.strip())
            else:
                pasos.append(ln)
        pasos_por_plato[plato_norm] = pasos
    return pasos_por_plato


def leer_compra_total(wb):
    ws = wb["Compra Total"]
    items = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or not row[1]:
            continue
        prov, cod, prod, udm, cant = row[1], row[2], row[3], row[4], row[5]
        if not prov or str(prov).startswith("Subtotal") or str(prov).startswith("TOTAL"):
            continue
        if not prod:
            continue
        items.append({
            "proveedor": str(prov), "codigo": str(cod or ""),
            "producto": str(prod), "udm": str(udm or ""),
            "cantidad": cant,
        })
    return items


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] No existe {EXCEL_PATH}")
        sys.exit(1)

    print(f"Leyendo {EXCEL_PATH}…")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)

    normas = leer_normas(wb)
    calendario, recetas_base = leer_menu(wb)
    pasos = leer_recetas_pasos(wb)
    compra_total = leer_compra_total(wb)

    # Inyectar pasos en recetas_base
    recetas = []
    for r in recetas_base:
        r["pasos"] = pasos.get(r["plato"], [])
        recetas.append(r)

    output = {
        "actualizado": date.today().isoformat(),
        "personas": 10,
        "normas": normas,
        "horarios": [
            {"nombre": "Mañana", "rango": "11:30 – 13:00"},
            {"nombre": "Tarde", "rango": "16:30 – 19:30"},
        ],
        "calendario": calendario,
        "recetas": recetas,
        "compraTotal": compra_total,
    }

    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"[OK] {JSON_PATH}")
    print(f"     Normas: {len(normas)}")
    print(f"     Días: {len(calendario)}")
    print(f"     Recetas: {len(recetas)} ({sum(1 for r in recetas if r['pasos'])} con pasos)")
    print(f"     Productos compra: {len(compra_total)}")


if __name__ == "__main__":
    main()
