"""
Genera el Excel maestro para subir a Google Sheets.
Pestañas: PRODUCTOS, PROVEEDORES, CONFIG
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import json
import os

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PRODUCTOS_JSON = os.path.join(SCRIPT_DIR, '..', 'PEDIDOS', 'APP', 'src', 'productos.json')
OUTPUT_PATH = os.path.join(SCRIPT_DIR, 'Casa Amparo 1948 - BBDD Maestro.xlsx')

# Load products
with open(PRODUCTOS_JSON, 'r', encoding='utf-8') as f:
    productos = json.load(f)

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2C1810', end_color='2C1810', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

accent_fill = PatternFill(start_color='8B4513', end_color='8B4513', fill_type='solid')
accent_font = Font(bold=True, color='FFFFFF', size=10)

stock_min_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
stock_min_font = Font(bold=True, color='E65100', size=11)

alt_fill = PatternFill(start_color='FAF7F3', end_color='FAF7F3', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='D5D5D5'),
    right=Side(style='thin', color='D5D5D5'),
    top=Side(style='thin', color='D5D5D5'),
    bottom=Side(style='thin', color='D5D5D5')
)

def style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = ws.dimensions

def style_data_rows(ws, max_row, num_cols):
    for row in range(2, max_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if row % 2 == 0:
                if col != 8:  # Don't override stock mínimo highlight
                    cell.fill = alt_fill

# ============================================================
# PESTAÑA: PRODUCTOS
# ============================================================
ws = wb.active
ws.title = 'PRODUCTOS'

headers = ['Proveedor', 'Categoría', 'Código', 'Producto', 'Precio', 'Unidad', 'IVA', 'Stock Mínimo']
for col, h in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=h)

# Deduplicate and sort
seen = set()
unique_productos = []
for p in productos:
    key = (p.get('proveedor', ''), p.get('codigo', ''), p.get('producto', ''))
    if key not in seen:
        seen.add(key)
        unique_productos.append(p)

# Sort by proveedor, then categoria, then producto
unique_productos.sort(key=lambda x: (x.get('proveedor', ''), x.get('categoria', ''), x.get('producto', '')))

for i, p in enumerate(unique_productos, 2):
    ws.cell(row=i, column=1, value=p.get('proveedor', ''))
    ws.cell(row=i, column=2, value=p.get('categoria', ''))
    ws.cell(row=i, column=3, value=p.get('codigo', ''))
    ws.cell(row=i, column=4, value=p.get('producto', ''))

    precio = p.get('precio')
    if isinstance(precio, (int, float)):
        cell = ws.cell(row=i, column=5, value=precio)
        cell.number_format = '#,##0.00 €'
    else:
        ws.cell(row=i, column=5, value=precio or '')

    ws.cell(row=i, column=6, value=p.get('unidad', ''))
    ws.cell(row=i, column=7, value=p.get('iva', ''))

    # Stock mínimo = 0 por defecto, highlighted for easy editing
    stock_cell = ws.cell(row=i, column=8, value=0)
    stock_cell.fill = stock_min_fill
    stock_cell.font = stock_min_font
    stock_cell.alignment = Alignment(horizontal='center', vertical='center')

max_row = len(unique_productos) + 1

style_header_row(ws, 8)
style_data_rows(ws, max_row, 8)

# Column widths
widths = {'A': 28, 'B': 24, 'C': 14, 'D': 48, 'E': 14, 'F': 14, 'G': 8, 'H': 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Freeze first row + first column
ws.freeze_panes = 'B2'

print(f'PRODUCTOS: {len(unique_productos)} productos')

# ============================================================
# PESTAÑA: PROVEEDORES
# ============================================================
ws_prov = wb.create_sheet('PROVEEDORES')

prov_headers = ['Nombre', 'Contacto', 'Teléfono', 'Email', 'Día Pedido', 'Pedido Mínimo', 'Notas']
for col, h in enumerate(prov_headers, 1):
    ws_prov.cell(row=1, column=col, value=h)

# Extract unique providers
proveedores_unicos = sorted(set(p.get('proveedor', '') for p in unique_productos if p.get('proveedor')))

# Known provider data
prov_data = {
    'DDI Nexia (Victoria/Damm)': {'contacto': 'Comercial DDI', 'dia': 'Lunes'},
    'Oido Cocina Gourmet': {'contacto': 'Pablo Salvador', 'telefono': '659 289 506', 'email': 'info@oidococinagourmet.com', 'dia': 'Jueves', 'minimo': '4 bolsas 2kg'},
    'Kemical': {'contacto': '', 'dia': ''},
    'Prodesco': {'contacto': '', 'dia': ''},
    'Garcimar': {'contacto': '', 'dia': ''},
    'Europastry': {'contacto': '', 'dia': ''},
    'Discarlux': {'contacto': '', 'dia': ''},
    'Oscar Casqueria': {'contacto': '', 'dia': ''},
}

for i, prov in enumerate(proveedores_unicos, 2):
    info = prov_data.get(prov, {})
    ws_prov.cell(row=i, column=1, value=prov)
    ws_prov.cell(row=i, column=2, value=info.get('contacto', ''))
    ws_prov.cell(row=i, column=3, value=info.get('telefono', ''))
    ws_prov.cell(row=i, column=4, value=info.get('email', ''))
    ws_prov.cell(row=i, column=5, value=info.get('dia', ''))
    ws_prov.cell(row=i, column=6, value=info.get('minimo', ''))
    ws_prov.cell(row=i, column=7, value=info.get('notas', ''))

style_header_row(ws_prov, 7)
style_data_rows(ws_prov, len(proveedores_unicos) + 1, 7)

prov_widths = {'A': 30, 'B': 22, 'C': 18, 'D': 32, 'E': 16, 'F': 20, 'G': 30}
for col, w in prov_widths.items():
    ws_prov.column_dimensions[col].width = w

ws_prov.freeze_panes = 'A2'

print(f'PROVEEDORES: {len(proveedores_unicos)} proveedores')

# ============================================================
# PESTAÑA: CONFIG
# ============================================================
ws_cfg = wb.create_sheet('CONFIG')

cfg_headers = ['Clave', 'Valor', 'Descripción']
for col, h in enumerate(cfg_headers, 1):
    ws_cfg.cell(row=1, column=col, value=h)

config_data = [
    ('nombre_negocio', 'Casa Amparo 1948', 'Nombre que aparece en PDFs y app'),
    ('subtitulo', 'Control de Inventario', 'Subtítulo de la app'),
    ('moneda', '€', 'Símbolo de moneda'),
    ('zona_horaria', 'Europe/Madrid', 'Zona horaria para registros'),
    ('version', '1.0', 'Versión de la BBDD'),
]

for i, (clave, valor, desc) in enumerate(config_data, 2):
    ws_cfg.cell(row=i, column=1, value=clave)
    ws_cfg.cell(row=i, column=2, value=valor)
    ws_cfg.cell(row=i, column=3, value=desc)

style_header_row(ws_cfg, 3)
style_data_rows(ws_cfg, len(config_data) + 1, 3)

cfg_widths = {'A': 22, 'B': 30, 'C': 45}
for col, w in cfg_widths.items():
    ws_cfg.column_dimensions[col].width = w

# ============================================================
# SAVE
# ============================================================
wb.save(OUTPUT_PATH)
print(f'\nExcel maestro guardado en: {OUTPUT_PATH}')
print('Ahora sube este archivo a Google Drive y conviértelo a Google Sheets.')
